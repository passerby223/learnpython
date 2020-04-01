# @Time    : 2020/3/28 下午11:27
# @File    : datatype.py
# @Author  : Always be codding
# @Software: PyCharm
import os
import timeit
from collections import namedtuple
from openpyxl import load_workbook  # 对已经存在的excel进行读写操作

"""
1.元组和列表
1.1 元组和列表性能比较分析
1.2 命名元组
2.字典和集合的原理和应用
2.1 散列类型的存储过程
2.2 字典查找的过程

3.推导式
3.1 列表推导式
3.2 字典推导式
3.3 生成器表达式

4.迭代器和生成器
4.1 生成器
4.2 迭代协议
4.3 迭代器
"""

''' 
1.1 元组和列表性能比较分析
'''

# def func():
#     for i in range(1, 4):
#         pass
#
#
# res1 = timeit.timeit('(1, 2, 3)')
# print('元组执行10000000次消耗{}秒'.format(res1))
# res = timeit.timeit('[1, 2, 3]')
# print('列表执行10000000次消耗{}秒'.format(res))
# res2 = timeit.timeit(func)
# print('方法执行10000000次消耗{}秒'.format(res2))

'''
1.2 命名元组 规则
'''
# name_tuple_pattern = namedtuple('name_tuple_info', ['name', 'age', 'gender'])
#
# my_name_tuple = name_tuple_pattern('哈哈哈', 21, gender='男')
# print(my_name_tuple)  # name_tuple_info(name='哈哈哈', age=21, gender='男')
# print(my_name_tuple.name)  # 哈哈哈
# print(my_name_tuple.age)  # 21
# print(my_name_tuple.gender)  # 男
'''
1.3 集合和字典
'''
# set1 = set()
# set2 = {1, 2, 3, 4, '哈哈'}
# dict1 = {"name": "张三", "age": 21}
# print(type(set1), type(set2), type(dict1)) # <class 'set'> <class 'set'> <class 'dict'>
# print(set1, set2, dict1) # set() {1, 2, 3, 4, '哈哈'} {'name': '张三', 'age': 21}
''' 
集合的方法
集合set是无序的，可以用来去重
'''
# set1.add('hello1')
# set1.add('hello2')
# set1.add('hello3')
# print(set1)
# set1.remove("hello1")
# print(set1)
# set1.add('hello4')
# set1.add('hello5')
# set1.pop() # 随机删除集合中的一个元素
# print(set1)
# set1.clear() # 清空集合中的所有元素
# print(set1) #　set()
# # 使用update(等同于list的extend方法)一次性向集合中添加多个元素
# set1.update(['111', "222", '333', '叭叭叭'])
# print(set1)
# set1.update(('444', "555", '666', '喇喇喇'))
# print(set1)
# set1.update({'777', "888", '999', '噜噜啦'})
# print(set1)

''' 
利用set集合中元素唯一的特性对list进行去重
'''
# list1 = [111, 111, 65, 89, 78, 89, 65, 78, "啦啦啦"]
# list1_new = list(set(list1))
# print("list1>>>{}, list1_new>>>{}".format(list1, list1_new))

''' 
2.字典和集合的原理和应用
2.1 散列类型的存储过程
2.2 字典查找的过程
Python中数据类型分类：
数值：1
序列类型：list,tuple
散列类型：dict,set

Python中数据类型可分为：可变类型和不可变类型(可hash的)

性能分析
查找元素时间：集合>字典>元组>列表
占用内存比较：字典>集合>列表>元组
'''

''' 
列表推导式
'''
# list_sheep = ['{}只羊'.format(i) for i in range(1, 11)]
# print('一共生成了{}只羊>>>{}'.format(len(list_sheep), list_sheep))

'''
字典推导式
'''
# dict_sheep = {str(i): '第{}只羊'.format(str(i)) for i in range(1, 11)}
# print('一共生成了{}只羊>>>{}'.format(len(dict_sheep), dict_sheep))

'''
小练习
将下边的字符串改成字典形式
'''
# cookie = 'UM_distinctid1=16f0634db1070;UM_distinctid2=066d9e63a3bcb4;UM_distinctid3=31760856;UM_distinctid4=1fa400'
# cookie_dict = {i.split('=')[0]: i.split('=')[1] for i in cookie.split(';')}
# print(cookie_dict)

''' 
生成器表达式
'''
# 如果最外层用`[]`的话，此表达式就是一个列表推导式
# 如果最外层用`{}`的话，此表达式就是一个字典推导式
# 如果最外层用`()`的话，会产生成一个生成器对象，次表达式称为生成器表达式
'''
思考
为什么要使用生成器？
定义一个生成器对象，不会占用很大的内存，它只是定义了一个生成器规则
不会像列表推导式、字典推导式那样直接生成一个list、dict存放在内存中，这样会很占用内存。
'''
# generator = (i for i in range(1, 11))
# print(generator)  # 生成器对象：<generator object <genexpr> at 0x7f6b2e033360>
# print(next(generator))  # 生成器对象的值可以通过next()方法依次取值
# print(next(generator))
# print(next(generator))
# print(next(generator))
# print(next(generator))
# print(generator.__next__())
# print(generator.__next__())
# print(generator.__next__())
# print(generator.__next__())
# print(generator.__next__())
# print(generator.__next__())  # 报错：StopIteration 停止迭代

''' 
yield 关键字的使用，yield关键字只能在函数中中使用

'''

# def func():
#     yield 1  # yield在函数中会返回一个生成器对象
#     print('哈哈哈')
#     yield '啦啦啦'
#
#
# res = func()
# print(res)  # <generator object func at 0x7f3088a533b8> 返回了一个生成器对象
# print(next(res)) # 1
# # 会将中间的print('哈哈哈') 的结果打印出来 # 哈哈哈
# print(res.__next__()) # 啦啦啦

''' 
1.迭代器协议

1.1迭代协议：一种是包含了`iter`方法的，另一种是包含了`getitem`方法的(比如str对象就没有`iter`方法，但是一样可以进行迭代操作)，
只要对象中包含了这两种方法的任意一种，那么这个对象就可以进行迭代操作。也就是实现了迭代协议。

1.2迭代器协议：
1.2.1实现了迭代器协议的对象(实现方式：对象内部定义了一个iter()方法)
1.2.2对象实现了__next__()方法
1.2.3__next__()方法返回了某个数值(当然一般情况下，我们需要的是返回这个对象的特定数字，并且按照一定的顺序进行依次返回。)
1.2.4__next__()方法需要在值取完时，抛出StopIteration错误信息

1.3可迭代对象
只要内部实现了迭代器协议就是一个可迭代对象(可迭代对象可以进行相关的迭代操作，比如for循环，map函数等)
所有的迭代器都是可迭代对象

2.迭代器
只要对象中实现了迭代协议，那么这就是一个迭代对象。如何将一个可迭代对象转换为一个迭代器？

'''
# list1 = [1, 2, 3, 4, 5]
# list2 = (1, 2, 3, 4, 5)
# list1_iterator = iter(list1)
# list2_iterator = iter(list2)
# print(list1_iterator)  # <list_iterator object at 0x7f1684a26ba8> list类型的迭代器对象
# print(list2_iterator)  # <tuple_iterator object at 0x7f6992c03908> tuple类型的迭代器对象

'''可迭代对象和迭代器的区别
可迭代对象：可以for循环遍历的都是可迭代对象，内部实现了__iter__()方法
迭代器：内部实现了__iter__()方法，还实现了__next__()方法
'''

'''生成器和迭代器的区别
生成器是迭代器的一种，如何区别迭代器和生成器？
1.生成器相比迭代器多了三个方法：send()、close()、throw()
send()：发送数据。可以与生成器交互进行数据传输
close()：关闭生成器
throw()：gen.throw(Exception,'Method throw called!')

生成器<迭代器<可迭代对象
'''

''' 
send()：发送数据。生成器对象的send()方法可以与生成器交互进行数据传输 代码示例
'''
# def gen():
#     for i in range(1, 6):
#         print('-' * 30)
#         res_gen = yield i
#         print('res_gen:', res_gen)
#
# res = gen()
# # print(res) # <generator object gen at 0x7f97cb4543b8> 生成器对象
# print(next(res)) #
# print(res.send('俺是生成器对象发送过来的数据!'))
# print(next(res)) #
# print(res.__next__()) #

''' 
会有以下输出：

------------------------------
1
res_gen: 俺是生成器对象发送过来的数据!
------------------------------
2
res_gen: None
------------------------------
3
res_gen: None
------------------------------
4
'''

''' 
通过分析上边的输入结果可知：
程序通过迭代获取第一次的数据时，先打印`------`，然后再返回yield的值1，然后程序暂停，然后再执行`print(res.send('俺是生成器对象发送过来的数据!'))`
然后程序回到之前的yield处，然后执行`print('res_gen:', res_gen)`......
'''

'''使用推导式对excel中的内容进行读取操作'''

from openpyxl import load_workbook  # 对已经存在的excel进行读写操作
import os


def read_data_from_excel(file_name, sheet_name):
    # 创建一个workbook对象
    wb = load_workbook(filename=file_name)
    ws = wb[sheet_name]
    column = ws.max_column + 1
    print('column', column)  # column 9
    row = ws.max_row + 1  # row 11
    print('row', row)
    data = ({ws.cell(1, j).value: ws.cell(i, j).value for j in range(1, column)} for i in range(2, row))
    return data


if __name__ == '__main__':
    current_dir_path = os.path.dirname(__file__)
    # print(current_dir_path)
    return_data = read_data_from_excel(file_name=current_dir_path + os.sep + 'data.xlsx',
                                       sheet_name='register')
    print(return_data)
    for i in return_data:
        print(i)

''' 
递归函数
'''
# def re_func():
#
